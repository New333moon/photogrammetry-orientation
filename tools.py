"""
本文件包含四个程序会用到的工具
包括xlsx数据读取、坐标转换和结果输出等
"""
import numpy as np
import csv
from pathlib import Path
from openpyxl import load_workbook

# 相机参数：
FOCAL_LENGTH = 92.0 # 主距mm
PIXEL_SIZE = 0.0039 # 像元大小mm
PHOTO_CENTER_I = 7295.5 # 像主点
PHOTO_CENTER_J = 12863.5
LEFT_PHOTO_NAME = "589" # 左右片编号，只用于输出说明
RIGHT_PHOTO_NAME = "590"


# 固定输入输出路径：
ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "inputs"
OUTPUT_DIR = ROOT / "outputs"

RELATIVE_XLSX = INPUT_DIR / "第4组汇总_相对定向同名点坐标.xlsx"
CONTROL_XLSX = INPUT_DIR / "第4组汇总_像控点同名坐标.xlsx"
OBJECT_XLSX = INPUT_DIR / "第4组汇总_待成图地物点同名坐标.xlsx"


# 保证输出文件夹存在
def ensure_output_dir():
    OUTPUT_DIR.mkdir(exist_ok=True)


# 像素坐标转像片坐标
def pixel_to_photo(i, j):
    x = (i - PHOTO_CENTER_I) * PIXEL_SIZE
    y = (PHOTO_CENTER_J - j) * PIXEL_SIZE
    return x, y


# 读取相对定向同名点
def read_relative_excel():# 返回一个列表，每个元素都是一个字典

    wb = load_workbook(RELATIVE_XLSX)
    ws = wb.active
    points = []

    # 前两行是表头，从第3行开始是数据
    for row in ws.iter_rows(min_row=3, values_only=True):
        point_id, li, lj, ri, rj = row[:5]
        if point_id is None:
            continue
        
        point = {
            "id": str(point_id),
            "left_i": float(li),
            "left_j": float(lj),
            "right_i": float(ri),
            "right_j": float(rj),
        }
        points.append(point)

    return points


# 读取像控点：左右像素坐标 + 地面坐标
def read_control_excel():
    
    wb = load_workbook(CONTROL_XLSX)
    ws = wb.active
    points = []

    # 前两行是表头，从第3行开始是数据
    for row in ws.iter_rows(min_row=3, values_only=True):
        point_id, number, li, lj, ri, rj, gx, gy, gz = row[:9]
        if point_id is None:
            continue
        
        point = {
                "id": str(point_id),
                "number": str(number),
                "left_i": float(li),
                "left_j": float(lj),
                "right_i": float(ri),
                "right_j": float(rj),
                "ground_x": float(gx),
                "ground_y": float(gy),
                "ground_z": float(gz),
        }
        points.append(point)
        
    return points


# 读取待成图地物点
def read_object_excel():

    wb = load_workbook(OBJECT_XLSX)
    ws = wb.active
    points = []
    current_name = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        object_name, number, li, lj, ri, rj = row[:6]
        if number is None:
            continue
        if object_name is not None:
            current_name = str(object_name)# 第一列为空时，沿用上一行的地物名称
       
        point = {
                "id": f"{current_name}_{number}",# 以地物名+序号来编号
                "object": current_name,
                "number": str(number),
                "left_i": float(li),
                "left_j": float(lj),
                "right_i": float(ri),
                "right_j": float(rj),
        }
        points.append(point)
        
    return points


# 用CSV保存结果
def write_csv(path, rows, fieldnames):
    ensure_output_dir()
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


# 读取前一步保存的CSV
def read_csv(path):
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


# 旋转矩阵
def rotation_matrix(omega, phi, kappa):
    # 计算三角函数
    co = np.cos(omega)
    so = np.sin(omega)
    cp = np.cos(phi)
    sp = np.sin(phi)
    ck = np.cos(kappa)
    sk = np.sin(kappa)

    # 绕x
    rx = np.array([
        [1, 0, 0],
        [0, co, -so],
        [0, so, co]
    ])
    # 绕y
    ry = np.array([
        [cp, 0, sp],
        [0, 1, 0],
        [-sp, 0, cp]
    ])
    # 绕z
    rz = np.array([
        [ck, -sk, 0],
        [sk, ck, 0],
        [0, 0, 1]
    ])

    R = rz @ ry @ rx

    return R


def image_ray(x, y, rotation):
    # 构造相机坐标系下的方向向量
    vec = np.array([x, y, -FOCAL_LENGTH])
    # 转到空间坐标系
    ray = rotation @ vec

    return ray
