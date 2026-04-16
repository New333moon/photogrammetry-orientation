"""
本文件放四个程序都会用到的小工具。

注意：
1. 相机参数直接固定写在这里，不从 inputs/camera 读取。
2. 四个主程序仍然是分开的：01/02/03/04 四个 py 文件。
3. 代码故意写得直白一些，便于实习报告说明。
"""

from __future__ import annotations

import csv
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


# =========================
# 固定相机参数：DMC3
# =========================

# 主距，单位 mm
FOCAL_LENGTH_MM = 92.0

# 像元大小。camera 文件中写的是 3.9 微米，换算为 0.0039 mm。
PIXEL_SIZE_MM = 0.0039

# 主点在像素坐标中的位置
PHOTO_CENTER_I = 7295.5
PHOTO_CENTER_J = 12863.5

# 本次实习使用的左右影像编号，只用于输出说明
LEFT_PHOTO_NAME = "589"
RIGHT_PHOTO_NAME = "590"


# =========================
# 固定输入输出路径
# =========================

ROOT = Path(__file__).resolve().parent
INPUT_DIR = ROOT / "inputs"
OUTPUT_DIR = ROOT / "outputs"

RELATIVE_XLSX = INPUT_DIR / "第4组汇总_相对定向同名点坐标.xlsx"
CONTROL_XLSX = INPUT_DIR / "第4组汇总_像控点同名坐标.xlsx"
OBJECT_XLSX = INPUT_DIR / "第4组汇总_待成图地物点同名坐标.xlsx"


def ensure_output_dir() -> None:
    """保证输出文件夹存在。"""
    OUTPUT_DIR.mkdir(exist_ok=True)


def pixel_to_photo(i: float, j: float) -> tuple[float, float]:
    """
    把像素坐标 (i, j) 转为像片坐标 (x, y)，单位 mm。

    数字影像的 i 向右增大，j 向下增大。
    摄影测量像片坐标一般让 x 向右、y 向上，所以 y 这里取反。
    """
    x = (float(i) - PHOTO_CENTER_I) * PIXEL_SIZE_MM
    y = (PHOTO_CENTER_J - float(j)) * PIXEL_SIZE_MM
    return x, y


def read_relative_excel() -> list[dict]:
    """读取相对定向同名点。"""
    wb = load_workbook(RELATIVE_XLSX, data_only=True)
    ws = wb.active
    points = []

    # 前两行是表头，从第 3 行开始是数据。
    for row in ws.iter_rows(min_row=3, values_only=True):
        point_id, li, lj, ri, rj = row[:5]
        if point_id is None:
            continue
        points.append(
            {
                "id": str(point_id),
                "left_i": float(li),
                "left_j": float(lj),
                "right_i": float(ri),
                "right_j": float(rj),
            }
        )
    return points


def read_control_excel() -> list[dict]:
    """读取像控点：左右像素坐标 + 地面坐标。"""
    wb = load_workbook(CONTROL_XLSX, data_only=True)
    ws = wb.active
    points = []

    for row in ws.iter_rows(min_row=3, values_only=True):
        point_id, number, li, lj, ri, rj, gx, gy, gz = row[:9]
        if point_id is None:
            continue
        points.append(
            {
                "id": str(point_id),
                "number": str(number),
                "left_i": float(li),
                "left_j": float(lj),
                "right_i": float(ri),
                "right_j": float(rj),
                "ground_x": float(gx),
                "ground_y": float(gy),
                "ground_z": float(gz),
            }
        )
    return points


def read_object_excel() -> list[dict]:
    """读取待成图地物点。第一列为空时，沿用上一行的地物名称。"""
    wb = load_workbook(OBJECT_XLSX, data_only=True)
    ws = wb.active
    points = []
    current_name = ""

    for row in ws.iter_rows(min_row=3, values_only=True):
        object_name, number, li, lj, ri, rj = row[:6]
        if number is None:
            continue
        if object_name is not None:
            current_name = str(object_name)
        point_id = f"{current_name}_{number}"
        points.append(
            {
                "id": point_id,
                "object": current_name,
                "number": str(number),
                "left_i": float(li),
                "left_j": float(lj),
                "right_i": float(ri),
                "right_j": float(rj),
            }
        )
    return points


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    """用 CSV 保存结果，Excel 也可以直接打开。"""
    ensure_output_dir()
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        # extrasaction="ignore" 表示：字典里多出来的字段不写入本 CSV。
        # 这样一个点在不同步骤中可以携带更多信息，输出时只挑需要的列。
        writer = csv.DictWriter(f, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def read_csv(path: Path) -> list[dict]:
    """读取前一步保存的 CSV。"""
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        return list(csv.DictReader(f))


def rotation_matrix(omega: float, phi: float, kappa: float) -> np.ndarray:
    """
    由三个角元素生成旋转矩阵。

    omega 绕 x 轴，phi 绕 y 轴，kappa 绕 z 轴。
    角度单位为弧度。
    """
    co, so = np.cos(omega), np.sin(omega)
    cp, sp = np.cos(phi), np.sin(phi)
    ck, sk = np.cos(kappa), np.sin(kappa)

    rx = np.array([[1, 0, 0], [0, co, -so], [0, so, co]], dtype=float)
    ry = np.array([[cp, 0, sp], [0, 1, 0], [-sp, 0, cp]], dtype=float)
    rz = np.array([[ck, -sk, 0], [sk, ck, 0], [0, 0, 1]], dtype=float)

    return rz @ ry @ rx


def image_ray(x: float, y: float, rotation: np.ndarray) -> np.ndarray:
    """把像点变成空间光线方向向量。"""
    return rotation @ np.array([x, y, -FOCAL_LENGTH_MM], dtype=float)
